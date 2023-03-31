import pytest
import openpyxl
from openpyxl import load_workbook
def get_data():
    wb=load_workbook(filename='c://testdata//Testdrives.xlsx')
    ws=wb.active
    data=[]
    for row in ws.iter_rows(min_row=1,max_col=2,values_only=True):

        data.append(row)
    return data
@pytest.mark.parametrize("input_1,expected_output",get_data())
def test_drives(input_1,expected_output):
    assert input_1==expected_output