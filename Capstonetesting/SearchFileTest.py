import pytest
import openpyxl
from openpyxl import load_workbook
def get_data():
    wb=load_workbook(filename='c://testdata//Testfiles.xlsx')
    ws=wb.active
    files=[]
    for row in ws.iter_rows(min_row=1,max_col=2,values_only=True):

        files.append(row)
    return files
@pytest.mark.parametrize("input_1,expected_output",get_data())
def test_files(input_1,expected_output):
    assert input_1==expected_output
