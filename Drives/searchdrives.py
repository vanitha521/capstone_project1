import openpyxl as xl
import os
"""To search drive so we are communicating with os """


class Searchdrives():
    """
    this module helps to find all drives present in pc
    67 to 90 for ASCII values
    chr() to convert ASCII to char
    """
    def __init__(self):  # consructor will help to initialiaze the var objects
        self.drives=[]
        self.workbook=xl.load_workbook("c://testdata//Testdrives.xlsx")

    def searchmydrives(self)->list:
        for x in range(67,91):
            if os.path.exists(chr(x)+":"):
                self.drives.append(chr(x))
        return self.drives


dr=Searchdrives()
data=str(dr.searchmydrives())
print(data)
sheet=dr.workbook.active
sheet.cell(row=1,column=1).value=data
sheet.cell(row=2,column=1).value=data
sheet.cell(row=3,column=1).value=data
dr.workbook.save("c://testdata//Testdrives.xlsx")
dr.workbook.close()


